from flask import Blueprint, request, jsonify, Response, send_file
from flask_jwt_extended import jwt_required, get_jwt_identity
from app.extensions import db
from app.models.sale import Sale
from app.models.sale_item import SaleItem
from app.models.product import Product
from app.models.monthly_snapshot import MonthlySnapshot
from app.models.user import User
from datetime import datetime, timedelta
from datetime import time as dt_time
from zoneinfo import ZoneInfo
import csv
import io
from openpyxl import Workbook

sale_bp = Blueprint("sales", __name__, url_prefix="/sales")
BUSINESS_TZ = ZoneInfo("America/Sao_Paulo")
UTC_TZ = ZoneInfo("UTC")


def _effective_unit_cost(func_obj):
    # Vendas antigas podem ter unit_cost=0; nesse caso usa custo atual do produto.
    return func_obj.coalesce(func_obj.nullif(SaleItem.unit_cost, 0), Product.cost, 0)


def _workspace_owner_id(current_user_id):
    user = User.query.get(current_user_id)
    if not user:
        return None
    return user.owner_id if user.owner_id else user.id


def _month_bounds(year, month):
    start = datetime(year, month, 1)
    if month == 12:
        end = datetime(year + 1, 1, 1)
    else:
        end = datetime(year, month + 1, 1)
    return start, end


def _business_day_utc_bounds(day_date):
    local_start = datetime.combine(day_date, dt_time.min, tzinfo=BUSINESS_TZ)
    local_end = local_start + timedelta(days=1)
    start_utc = local_start.astimezone(UTC_TZ).replace(tzinfo=None)
    end_utc = local_end.astimezone(UTC_TZ).replace(tzinfo=None)
    return start_utc, end_utc


def _business_month_utc_bounds(year, month):
    local_start = datetime(year, month, 1, tzinfo=BUSINESS_TZ)
    if month == 12:
        local_end = datetime(year + 1, 1, 1, tzinfo=BUSINESS_TZ)
    else:
        local_end = datetime(year, month + 1, 1, tzinfo=BUSINESS_TZ)
    start_utc = local_start.astimezone(UTC_TZ).replace(tzinfo=None)
    end_utc = local_end.astimezone(UTC_TZ).replace(tzinfo=None)
    return start_utc, end_utc


def _monthly_summary(user_id, year, month):
    start, end = _business_month_utc_bounds(year, month)

    sales = Sale.query.filter(
        Sale.user_id == user_id,
        Sale.created_at >= start,
        Sale.created_at < end
    ).order_by(Sale.created_at.asc()).all()

    total_sales = len(sales)
    from sqlalchemy import func

    gross_amount_result = db.session.query(
        func.sum(Sale.total)
    ).filter(
        Sale.user_id == user_id,
        Sale.created_at >= start,
        Sale.created_at < end
    ).scalar()

    total_profit_result = db.session.query(
        func.sum((SaleItem.price - _effective_unit_cost(func)) * SaleItem.quantity)
    ).join(
        Sale, Sale.id == SaleItem.sale_id
    ).join(
        Product, Product.id == SaleItem.product_id
    ).filter(
        Sale.user_id == user_id,
        Sale.created_at >= start,
        Sale.created_at < end
    ).scalar()

    total_amount = float(total_profit_result) if total_profit_result else 0.0
    gross_amount = float(gross_amount_result) if gross_amount_result else 0.0
    average_sale = (total_amount / total_sales) if total_sales else 0.0

    top_products = db.session.query(
        Product.name,
        func.sum(SaleItem.quantity).label('total_quantity'),
        func.sum((SaleItem.price - _effective_unit_cost(func)) * SaleItem.quantity).label('total_amount')
    ).join(
        SaleItem, SaleItem.product_id == Product.id
    ).join(
        Sale, Sale.id == SaleItem.sale_id
    ).filter(
        Sale.user_id == user_id,
        Sale.created_at >= start,
        Sale.created_at < end
    ).group_by(
        Product.id, Product.name
    ).order_by(
        func.sum(SaleItem.quantity).desc()
    ).limit(5).all()

    return {
        "period": {
            "year": year,
            "month": month,
            "start_date": datetime(year, month, 1).date().isoformat(),
            "end_date": (_month_bounds(year, month)[1] - timedelta(days=1)).date().isoformat()
        },
        "summary": {
            "total_sales": total_sales,
            "total_amount": total_amount,
            "gross_amount": gross_amount,
            "total_profit": total_amount,
            "average_sale": average_sale
        },
        "sales": sales,
        "top_products": [
            {
                "name": name,
                "total_quantity": int(total_quantity),
                "total_amount": float(total_amount)
            }
            for name, total_quantity, total_amount in top_products
        ]
    }


# ======================
# LISTAR VENDAS
# ======================
@sale_bp.route("/", methods=["GET"])
@jwt_required()
def list_sales():
    try:
        current_user_id = int(get_jwt_identity())
        user_id = _workspace_owner_id(current_user_id)
        if not user_id:
            return jsonify({"error": "Usuário não encontrado"}), 404
        
        # Filtros
        start_date = request.args.get("start_date")
        end_date = request.args.get("end_date")
        min_total = request.args.get("min_total")
        max_total = request.args.get("max_total")
        
        query = Sale.query.filter_by(user_id=user_id)
        
        # Aplicar filtros de data
        if start_date:
            try:
                start = datetime.fromisoformat(start_date.replace('Z', '+00:00'))
                query = query.filter(Sale.created_at >= start)
            except ValueError:
                return jsonify({"error": "Data inicial inválida"}), 400
        
        if end_date:
            try:
                end = datetime.fromisoformat(end_date.replace('Z', '+00:00'))
                # Adicionar 1 dia para incluir todo o dia final
                end = end + timedelta(days=1)
                query = query.filter(Sale.created_at < end)
            except ValueError:
                return jsonify({"error": "Data final inválida"}), 400
        
        # Aplicar filtros de total
        if min_total:
            try:
                query = query.filter(Sale.total >= float(min_total))
            except ValueError:
                return jsonify({"error": "Total mínimo inválido"}), 400
        
        if max_total:
            try:
                query = query.filter(Sale.total <= float(max_total))
            except ValueError:
                return jsonify({"error": "Total máximo inválido"}), 400
        
        # Ordenação
        query = query.order_by(Sale.created_at.desc())
        
        # Paginação
        page = int(request.args.get("page", 1))
        per_page = int(request.args.get("per_page", 20))
        
        total = query.count()
        sales = query.offset((page - 1) * per_page).limit(per_page).all()
        
        return jsonify({
            "sales": [s.to_dict() for s in sales],
            "pagination": {
                "page": page,
                "per_page": per_page,
                "total": total,
                "pages": (total + per_page - 1) // per_page
            },
            "summary": {
                "total_sales": total,
                "total_amount": sum(float(s.total) for s in sales),
                "average_sale": sum(float(s.total) for s in sales) / len(sales) if sales else 0
            }
        }), 200
        
    except Exception as e:
        return jsonify({"error": f"Erro ao listar vendas: {str(e)}"}), 500


# ======================
# OBTER VENDA POR ID
# ======================
@sale_bp.route("/<int:sale_id>", methods=["GET"])
@jwt_required()
def get_sale(sale_id):
    try:
        current_user_id = int(get_jwt_identity())
        user_id = _workspace_owner_id(current_user_id)
        if not user_id:
            return jsonify({"error": "Usuário não encontrado"}), 404
        
        sale = Sale.query.filter_by(id=sale_id, user_id=user_id).first()
        
        if not sale:
            return jsonify({"error": "Venda não encontrada"}), 404
        
        return jsonify(sale.to_dict(include_items=True)), 200
        
    except Exception as e:
        return jsonify({"error": f"Erro ao obter venda: {str(e)}"}), 500


# ======================
# CRIAR VENDA
# ======================
@sale_bp.route("/", methods=["POST"])
@jwt_required()
def create_sale():
    try:
        current_user_id = int(get_jwt_identity())
        user_id = _workspace_owner_id(current_user_id)
        if not user_id:
            return jsonify({"error": "Usuário não encontrado"}), 404
        data = request.get_json()
        
        if not data:
            return jsonify({"error": "Dados não fornecidos"}), 400

        items = data.get("items")
        
        if not items or not isinstance(items, list):
            return jsonify({"error": "Itens da venda são obrigatórios e devem ser uma lista"}), 400
        
        if len(items) == 0:
            return jsonify({"error": "A venda deve conter pelo menos um item"}), 400

        total = 0
        sale_items = []
        
        # 🔎 Validação ANTECIPADA de todos os itens
        for i, item in enumerate(items):
            product_id = item.get("product_id")
            quantity = item.get("quantity")
            
            if not product_id or not quantity:
                return jsonify({"error": f"Item {i+1}: produto_id e quantidade são obrigatórios"}), 400
            
            try:
                quantity = int(quantity)
                if quantity <= 0:
                    return jsonify({"error": f"Item {i+1}: quantidade deve ser maior que zero"}), 400
            except (ValueError, TypeError):
                return jsonify({"error": f"Item {i+1}: quantidade deve ser um número inteiro"}), 400
            
            # Buscar produto
            product = Product.query.filter_by(
                id=product_id,
                user_id=user_id
            ).with_for_update().first()  # Lock para evitar race condition
            
            if not product:
                return jsonify({"error": f"Item {i+1}: produto não encontrado"}), 404
            
            # Validar estoque
            if product.stock < quantity:
                return jsonify({
                    "error": f"Item {i+1}: estoque insuficiente para '{product.name}'. Disponível: {product.stock}"
                }), 400
            
            subtotal = float(product.price) * quantity
            total += subtotal
            
            sale_items.append({
                "product": product,
                "quantity": quantity,
                "price": float(product.price),
                "unit_cost": float(product.cost),
                "subtotal": subtotal
            })
        
        # 🧾 Criar venda em uma transação atômica
        try:
            db.session.begin_nested()
            
            # Criar venda
            sale = Sale(total=total, user_id=user_id, employee_id=current_user_id)
            db.session.add(sale)
            db.session.flush()  # gera ID da venda sem commit
            
            # 📦 Criar itens + baixar estoque
            for item in sale_items:
                sale_item = SaleItem(
                    sale_id=sale.id,
                    product_id=item["product"].id,
                    quantity=item["quantity"],
                    price=item["price"],
                    unit_cost=item["unit_cost"]
                )
                
                # Atualizar estoque
                item["product"].stock -= item["quantity"]
                
                db.session.add(sale_item)
            
            db.session.commit()
            
            # Buscar venda completa para resposta
            sale_complete = Sale.query.get(sale.id)
            
            return jsonify({
                "message": "Venda registrada com sucesso",
                "sale": sale_complete.to_dict(include_items=True)
            }), 201
            
        except Exception as e:
            db.session.rollback()
            raise e
        
    except Exception as e:
        return jsonify({"error": f"Erro ao processar venda: {str(e)}"}), 500


# ======================
# ESTATÍSTICAS DE VENDAS
# ======================
@sale_bp.route("/stats", methods=["GET"])
@jwt_required()
def get_stats():
    try:
        current_user_id = int(get_jwt_identity())
        user_id = _workspace_owner_id(current_user_id)
        if not user_id:
            return jsonify({"error": "Usuário não encontrado"}), 404
        
        # Filtros de período
        days = int(request.args.get("days", 30))
        reference_date = request.args.get("reference_date")

        if reference_date:
            try:
                base_day_date = datetime.strptime(reference_date, "%Y-%m-%d").date()
            except ValueError:
                return jsonify({"error": "reference_date inválida. Use YYYY-MM-DD."}), 400
        else:
            base_day_date = datetime.now().date()

        base_day_start, base_day_end = _business_day_utc_bounds(base_day_date)
        
        # Data de início
        start_date = base_day_start - timedelta(days=max(days - 1, 0))
        
        # Total de vendas no período
        total_sales = Sale.query.filter(
            Sale.user_id == user_id,
            Sale.created_at >= start_date,
            Sale.created_at < base_day_end
        ).count()
        
        # Valor bruto vendido
        gross_amount_result = db.session.query(
            db.func.sum(Sale.total)
        ).filter(
            Sale.user_id == user_id,
            Sale.created_at >= start_date,
            Sale.created_at < base_day_end
        ).scalar()

        # Total liquido (valor de venda - custo dos itens)
        total_profit_result = db.session.query(
            db.func.sum((SaleItem.price - _effective_unit_cost(db.func)) * SaleItem.quantity)
        ).join(
            Sale, Sale.id == SaleItem.sale_id
        ).join(
            Product, Product.id == SaleItem.product_id
        ).filter(
            Sale.user_id == user_id,
            Sale.created_at >= start_date,
            Sale.created_at < base_day_end
        ).scalar()

        gross_amount = float(gross_amount_result) if gross_amount_result else 0.0
        total_amount = float(total_profit_result) if total_profit_result else 0.0
        
        # Vendas por dia (últimos 7 dias)
        daily_sales = []
        today_net = 0.0
        today_gross = 0.0
        today_date = None
        today_sales_count = 0
        for i in range(7):
            day_local = base_day_date - timedelta(days=i)
            day_start, day_end = _business_day_utc_bounds(day_local)

            day_count = Sale.query.filter(
                Sale.user_id == user_id,
                Sale.created_at >= day_start,
                Sale.created_at < day_end
            ).count()

            day_gross = db.session.query(
                db.func.sum(Sale.total)
            ).filter(
                Sale.user_id == user_id,
                Sale.created_at >= day_start,
                Sale.created_at < day_end
            ).scalar()
            
            day_total = db.session.query(
                db.func.sum((SaleItem.price - _effective_unit_cost(db.func)) * SaleItem.quantity)
            ).join(
                Sale, Sale.id == SaleItem.sale_id
            ).join(
                Product, Product.id == SaleItem.product_id
            ).filter(
                Sale.user_id == user_id,
                Sale.created_at >= day_start,
                Sale.created_at < day_end
            ).scalar()
            
            daily_sales.append({
                "date": day_local.isoformat(),
                "sales_count": int(day_count),
                "total": float(day_total) if day_total else 0.0,
                "total_profit": float(day_total) if day_total else 0.0,
                "gross_total": float(day_gross) if day_gross else 0.0
            })

            if i == 0:
                today_date = day_start.date().isoformat()
                today_net = float(day_total) if day_total else 0.0
                today_gross = float(day_gross) if day_gross else 0.0
                today_sales_count = int(day_count)
        
        # Produtos mais vendidos
        from sqlalchemy import func
        
        top_products = db.session.query(
            Product.name,
            func.sum(SaleItem.quantity).label('total_quantity'),
            func.sum((SaleItem.price - _effective_unit_cost(func)) * SaleItem.quantity).label('total_amount')
        ).join(
            SaleItem, SaleItem.product_id == Product.id
        ).join(
            Sale, Sale.id == SaleItem.sale_id
        ).filter(
            Sale.user_id == user_id,
            Sale.created_at >= start_date
        ).group_by(
            Product.id, Product.name
        ).order_by(
            func.sum(SaleItem.quantity).desc()
        ).limit(5).all()
        
        return jsonify({
            "period": {
                "days": days,
                "start_date": start_date.isoformat()
            },
            "summary": {
                "total_sales": total_sales,
                "total_amount": total_amount,
                "gross_amount": gross_amount,
                "total_profit": total_amount,
                "today_date": today_date,
                "today_net": today_net,
                "today_gross": today_gross,
                "today_sales_count": today_sales_count,
                "average_sale": total_amount / total_sales if total_sales > 0 else 0
            },
            "daily_sales": daily_sales,
            "top_products": [
                {
                    "name": name,
                    "total_quantity": int(total_quantity),
                    "total_amount": float(total_amount)
                }
                for name, total_quantity, total_amount in top_products
            ]
        }), 200
        
    except Exception as e:
        return jsonify({"error": f"Erro ao obter estatísticas: {str(e)}"}), 500


@sale_bp.route("/reports/monthly", methods=["GET"])
@jwt_required()
def get_monthly_report():
    try:
        current_user_id = int(get_jwt_identity())
        user_id = _workspace_owner_id(current_user_id)
        if not user_id:
            return jsonify({"error": "Usuário não encontrado"}), 404
        now = datetime.utcnow()

        year = int(request.args.get("year", now.year))
        month = int(request.args.get("month", now.month))
        report_format = request.args.get("format", "json").lower()

        if month < 1 or month > 12:
            return jsonify({"error": "Mês inválido. Use um valor entre 1 e 12."}), 400

        report = _monthly_summary(user_id, year, month)

        if report_format == "json":
            return jsonify({
                "period": report["period"],
                "summary": report["summary"],
                "top_products": report["top_products"],
                "sales": [sale.to_dict(include_items=True) for sale in report["sales"]]
            }), 200

        if report_format == "csv":
            output = io.StringIO()
            writer = csv.writer(output)

            writer.writerow(["Ano", "Mes", "Total Vendas", "Total Faturado", "Ticket Medio"])
            writer.writerow([
                report["period"]["year"],
                report["period"]["month"],
                report["summary"]["total_sales"],
                f"{report['summary']['total_amount']:.2f}",
                f"{report['summary']['average_sale']:.2f}"
            ])
            writer.writerow([])
            writer.writerow(["Venda ID", "Data", "Itens", "Total"])

            for sale in report["sales"]:
                writer.writerow([
                    sale.id,
                    sale.created_at.isoformat() if sale.created_at else "",
                    len(sale.items),
                    f"{float(sale.total):.2f}"
                ])

            filename = f"relatorio-mensal-{year}-{month:02d}.csv"
            return Response(
                output.getvalue(),
                mimetype="text/csv",
                headers={"Content-Disposition": f"attachment; filename={filename}"}
            )

        if report_format == "xlsx":
            wb = Workbook()

            ws_summary = wb.active
            ws_summary.title = "Resumo"
            ws_summary.append(["Ano", "Mes", "Total Vendas", "Total Faturado", "Ticket Medio"])
            ws_summary.append([
                report["period"]["year"],
                report["period"]["month"],
                report["summary"]["total_sales"],
                report["summary"]["total_amount"],
                report["summary"]["average_sale"]
            ])

            ws_sales = wb.create_sheet("Vendas")
            ws_sales.append(["Venda ID", "Data", "Itens", "Total"])
            for sale in report["sales"]:
                ws_sales.append([
                    sale.id,
                    sale.created_at.isoformat() if sale.created_at else "",
                    len(sale.items),
                    float(sale.total)
                ])

            ws_products = wb.create_sheet("Top Produtos")
            ws_products.append(["Produto", "Quantidade", "Total"])
            for product in report["top_products"]:
                ws_products.append([
                    product["name"],
                    product["total_quantity"],
                    product["total_amount"]
                ])

            stream = io.BytesIO()
            wb.save(stream)
            stream.seek(0)

            filename = f"relatorio-mensal-{year}-{month:02d}.xlsx"
            return send_file(
                stream,
                as_attachment=True,
                download_name=filename,
                mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

        return jsonify({"error": "Formato inválido. Use json, csv ou xlsx."}), 400

    except ValueError:
        return jsonify({"error": "Parâmetros inválidos para ano/mês."}), 400
    except Exception as e:
        return jsonify({"error": f"Erro ao gerar relatório mensal: {str(e)}"}), 500


@sale_bp.route("/reports/monthly/compare", methods=["GET"])
@jwt_required()
def compare_monthly_report():
    try:
        current_user_id = int(get_jwt_identity())
        user_id = _workspace_owner_id(current_user_id)
        if not user_id:
            return jsonify({"error": "Usuário não encontrado"}), 404
        now = datetime.utcnow()

        year = int(request.args.get("year", now.year))
        month = int(request.args.get("month", now.month))

        if month < 1 or month > 12:
            return jsonify({"error": "Mês inválido. Use um valor entre 1 e 12."}), 400

        current_report = _monthly_summary(user_id, year, month)

        if month == 1:
            prev_year = year - 1
            prev_month = 12
        else:
            prev_year = year
            prev_month = month - 1

        previous_report = _monthly_summary(user_id, prev_year, prev_month)

        current_total = current_report["summary"]["total_amount"]
        previous_total = previous_report["summary"]["total_amount"]

        if previous_total == 0:
            growth_percent = 100.0 if current_total > 0 else 0.0
        else:
            growth_percent = ((current_total - previous_total) / previous_total) * 100

        current_count = current_report["summary"]["total_sales"]
        previous_count = previous_report["summary"]["total_sales"]

        return jsonify({
            "current": {
                "year": year,
                "month": month,
                "summary": current_report["summary"]
            },
            "previous": {
                "year": prev_year,
                "month": prev_month,
                "summary": previous_report["summary"]
            },
            "comparison": {
                "amount_difference": current_total - previous_total,
                "sales_difference": current_count - previous_count,
                "growth_percent": growth_percent
            }
        }), 200

    except ValueError:
        return jsonify({"error": "Parâmetros inválidos para ano/mês."}), 400
    except Exception as e:
        return jsonify({"error": f"Erro ao comparar meses: {str(e)}"}), 500


@sale_bp.route("/reports/monthly/save", methods=["POST"])
@jwt_required()
def save_monthly_snapshot():
    try:
        current_user_id = int(get_jwt_identity())
        user_id = _workspace_owner_id(current_user_id)
        if not user_id:
            return jsonify({"error": "Usuário não encontrado"}), 404
        now = datetime.utcnow()

        data = request.get_json(silent=True) or {}
        year = int(data.get("year", now.year))
        month = int(data.get("month", now.month))

        if month < 1 or month > 12:
            return jsonify({"error": "Mês inválido. Use um valor entre 1 e 12."}), 400

        report = _monthly_summary(user_id, year, month)
        summary = report["summary"]

        snapshot = MonthlySnapshot.query.filter_by(user_id=user_id, year=year, month=month).first()
        if not snapshot:
            snapshot = MonthlySnapshot(user_id=user_id, year=year, month=month)
            db.session.add(snapshot)

        snapshot.total_sales = int(summary.get("total_sales", 0) or 0)
        snapshot.total_amount = float(summary.get("total_amount", 0) or 0)
        snapshot.gross_amount = float(summary.get("gross_amount", 0) or 0)
        snapshot.saved_at = datetime.utcnow()

        db.session.commit()

        return jsonify({
            "message": "Resumo mensal salvo com sucesso",
            "snapshot": snapshot.to_dict(),
        }), 200

    except ValueError:
        return jsonify({"error": "Parâmetros inválidos para ano/mês."}), 400
    except Exception as e:
        db.session.rollback()
        return jsonify({"error": f"Erro ao salvar resumo mensal: {str(e)}"}), 500


@sale_bp.route("/reports/monthly/saved", methods=["GET"])
@jwt_required()
def list_monthly_snapshots():
    try:
        current_user_id = int(get_jwt_identity())
        user_id = _workspace_owner_id(current_user_id)
        if not user_id:
            return jsonify({"error": "Usuário não encontrado"}), 404

        snapshots = MonthlySnapshot.query.filter_by(user_id=user_id).order_by(
            MonthlySnapshot.year.desc(),
            MonthlySnapshot.month.desc()
        ).all()

        return jsonify({"snapshots": [snapshot.to_dict() for snapshot in snapshots]}), 200

    except Exception as e:
        return jsonify({"error": f"Erro ao listar meses salvos: {str(e)}"}), 500


@sale_bp.route("/reports/monthly/saved/<int:year>/<int:month>", methods=["GET"])
@jwt_required()
def get_saved_monthly_snapshot(year, month):
    try:
        current_user_id = int(get_jwt_identity())
        user_id = _workspace_owner_id(current_user_id)
        if not user_id:
            return jsonify({"error": "Usuário não encontrado"}), 404

        if month < 1 or month > 12:
            return jsonify({"error": "Mês inválido. Use um valor entre 1 e 12."}), 400

        snapshot = MonthlySnapshot.query.filter_by(user_id=user_id, year=year, month=month).first()

        if not snapshot:
            return jsonify({"error": "Resumo mensal salvo não encontrado."}), 404

        return jsonify({"snapshot": snapshot.to_dict()}), 200

    except Exception as e:
        return jsonify({"error": f"Erro ao obter mês salvo: {str(e)}"}), 500